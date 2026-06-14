"""CSV, Excel, and OSCAL export for POAM findings."""

import csv
from pathlib import Path

from .models import Finding

CSV_HEADERS = [
    "Weakness Name",
    "Weakness Description",
    "Security Control",
    "Severity",
    "Source",
    "Detection Date",
    "Scheduled Completion",
    "Office/Org",
    "Point of Contact",
    "Resources Required",
    "Remediation Plan",
    "Milestones",
    "Status",
    "Comments",
]


def write_csv(findings: list[Finding], output_path: str | Path) -> None:
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        for finding in findings:
            writer.writerow(finding.to_csv_row())


def write_excel(findings: list[Finding], output_path: str | Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for Excel export: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POAM"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(CSV_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap

    severity_fills = {
        "Critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "High": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        "Medium": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "Low": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    }
    severity_col = CSV_HEADERS.index("Severity") + 1

    for row_idx, finding in enumerate(findings, start=2):
        row_data = finding.to_csv_row()
        for col_idx, header in enumerate(CSV_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.alignment = wrap

        sev = row_data["Severity"]
        if sev in severity_fills:
            sev_cell = ws.cell(row=row_idx, column=severity_col)
            sev_cell.fill = severity_fills[sev]
            sev_cell.font = Font(bold=True)

    col_widths = [30, 50, 18, 12, 20, 16, 20, 30, 25, 30, 50, 50, 15, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)
